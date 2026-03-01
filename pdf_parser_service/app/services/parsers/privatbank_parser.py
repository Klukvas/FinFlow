from typing import List, Optional
from openpyxl import load_workbook
from app.models.transaction import ParsedTransaction, BankType, TransactionType
from .base_parser import BasePDFParser


class PrivatbankParser(BasePDFParser):
    """Parser for PrivatBank XLSX statements"""

    def __init__(self):
        super().__init__(BankType.PRIVATBANK)

    async def parse_pdf(
        self, file_path: str, language: str = "en", user_id: int = None
    ) -> List[ParsedTransaction]:
        """Parse PrivatBank XLSX file and extract transactions"""
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            ws = wb.active

            transactions: List[ParsedTransaction] = []

            # Row 1 = title, Row 2 = headers, Row 3+ = data
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                # Stop at first completely empty row
                if row is None or all(cell is None for cell in row):
                    break

                try:
                    transaction = self._parse_xlsx_row(row, row_idx)
                    if transaction:
                        transactions.append(transaction)
                except Exception as e:
                    self.logger.warning(
                        f"Failed to parse PrivatBank row {row_idx}: {e}"
                    )
                    continue

            self.logger.info(
                f"Successfully parsed {len(transactions)} PrivatBank transactions"
            )
            return transactions

        except Exception as e:
            self.logger.error(f"Error parsing PrivatBank XLSX: {e}")
            raise
        finally:
            wb.close()

    async def _extract_transactions_from_table(
        self, table: List[List[str]], language: str = "en", user_id: int = None
    ) -> List[ParsedTransaction]:
        """Not used for XLSX flow — required by abstract base class"""
        return []

    def _parse_xlsx_row(
        self, row: tuple, row_idx: int
    ) -> Optional[ParsedTransaction]:
        """Parse a single PrivatBank XLSX row.

        Column layout:
        0: Date+time (str DD.MM.YYYY HH:MM:SS)
        1: Category (str)
        2: Card (str)
        3: Description (str)
        4: Amount in card currency (float, sign determines income/expense)
        5: Card currency (str)
        6: Amount in tx currency (float)
        7: Tx currency (str)
        8: Balance after (float)
        9: Balance currency (str)
        """
        if len(row) < 6:
            return None

        # --- date ---
        date_str = str(row[0]).strip() if row[0] else ""
        transaction_date = self._parse_date(date_str)
        if transaction_date is None:
            self.logger.warning(
                f"Row {row_idx}: could not parse date '{date_str}'"
            )
            return None

        # --- amount ---
        raw_amount = row[4]
        if raw_amount is None:
            return None

        try:
            amount = float(raw_amount)
        except (ValueError, TypeError):
            self.logger.warning(
                f"Row {row_idx}: could not parse amount '{raw_amount}'"
            )
            return None

        if amount == 0:
            return None

        # --- transaction type ---
        transaction_type = (
            TransactionType.INCOME if amount > 0 else TransactionType.EXPENSE
        )

        # --- description ---
        category = str(row[1]).strip() if row[1] else ""
        description_raw = str(row[3]).strip() if row[3] else ""

        if category and description_raw:
            description = f"[{category}] {description_raw}"
        elif description_raw:
            description = description_raw
        elif category:
            description = category
        else:
            description = "PrivatBank transaction"

        description = self._clean_description(description)

        # --- currency ---
        currency = str(row[5]).strip() if row[5] else "UAH"

        # --- confidence ---
        confidence_score = self._calculate_confidence_score(
            description, transaction_date, amount
        )

        # --- raw text for debugging ---
        raw_text = (
            f"{date_str} | {category} | {description_raw} | {raw_amount} {currency}"
        )

        return ParsedTransaction(
            amount=abs(amount),
            description=description,
            transaction_date=transaction_date,
            transaction_type=transaction_type,
            bank_type=self.bank_type,
            raw_text=raw_text,
            confidence_score=confidence_score,
            mcc_code=None,
            mcc_category_name=None,
            mcc_category_translation=None,
            category_exists=False,
            category_id=None,
        )
